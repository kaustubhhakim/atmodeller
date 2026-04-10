# SPDX-FileCopyrightText: 2024 Dan J. Bower <dbower@eaps.ethz.ch>
#
# SPDX-License-Identifier: GPL-3.0-or-later

"""Public output API for atmodeller results.

This module exposes the high-level output interface used to inspect, compare, and export simulation
results. It builds on base output infrastructure defined in :mod:`atmodeller.output_base` and
provides user-facing convenience methods for:

- Producing nested dictionary outputs in multiple formats
- Converting outputs to pandas DataFrames
- Exporting outputs to Excel and pickle
- Comparing outputs against reference data and quick logging/inspection

The :class:`Output` class is the main entry point for downstream workflows.

.. warning::
    Methods that convert to NumPy, compare values, or perform file export are intended for
    post-processing and are not compatible with JAX-compiled execution contexts (for example,
    inside ``jax.jit``).
"""

import logging
import pickle
from pathlib import Path
from pprint import pformat
from typing import Any, Literal

import equinox as eqx
import numpy as np
import pandas as pd
from jaxtyping import Array, Float
from openpyxl.styles import PatternFill

from atmodeller import __version__
from atmodeller.containers import MultiAttemptSolution
from atmodeller.jax_utils import NpArray
from atmodeller.output_base import (
    OutputElementsSpeciesDict,
    OutputNamedArraysDict,
    OutputNaturalDict,
    flatten_dictionary,
)
from atmodeller.parameters import Parameters

logger: logging.Logger = logging.getLogger(__name__)


class Output(eqx.Module):
    """Master output interface for model results.

    This class provides a unified entry point for all output formats, conversions, comparisons,
    and exports. It wraps the various output dictionary representations (natural, named arrays,
    grouped by element/species) and exposes methods for:

    - Outputting results in different formats (dict, DataFrame, Excel, pickle)
    - Comparing outputs for regression testing/validation
    - Quick inspection and logging

    .. warning::
        Some methods (such as ``to_dict(to_numpy=True)``, ``compare``, and export methods) are
        not compatible with JAX-compiled workflows (e.g., inside a ``jax.jit`` context), as they
        may use operations or objects that are not supported by JAX transformations.

    Args:
        parameters: Parameters
        multi_attempt_solution: Multiple attempt solution object
    """

    parameters: Parameters
    multi_attempt_solution: MultiAttemptSolution

    @property
    def solution(self) -> Float[Array, "#n_batch twice_species"]:
        """Solution array for all species i.e. log number of moles and log stability"""
        return self.multi_attempt_solution.value

    def to_dict(
        self,
        output_format: Literal["natural", "named_arrays", "elements_species"] = "named_arrays",
        to_numpy: bool = False,
        **kwargs,
    ) -> dict[str, Any]:
        """Output as a nested dictionary with JAX or NumPy arrays.

        .. warning::
            ``to_numpy`` must be ``False`` if used within a jitted context, as NumPy arrays are not
            compatible with JAX transformations (jit, vmap, etc.).

        Args:
            output_format: The format of the output dictionary. Can be ``natural`` for the natural
                output format based on the arrays used internally, ``named_arrays`` for an
                alternative format with named arrays, or ``elements_species`` for an alternative
                format grouped by element and species names. Defaults to ``named_arrays``.
            to_numpy: Whether to convert JAX arrays to NumPy arrays. Defaults to ``False``.
                Must be ``False`` if used within a jitted context, as NumPy arrays are not
                compatible with JAX transformations (jit, vmap, etc.).
            **kwargs: Arbitrary keyword arguments for the output dictionary

        Returns:
            Dictionary of the solution with JAX or NumPy arrays in the specified format
        """
        if output_format == "natural":
            return OutputNaturalDict(self.parameters, self.multi_attempt_solution).to_dict(
                to_numpy=to_numpy, **kwargs
            )
        elif output_format == "named_arrays":
            return OutputNamedArraysDict(self.parameters, self.multi_attempt_solution).to_dict(
                to_numpy=to_numpy, **kwargs
            )
        elif output_format == "elements_species":
            return OutputElementsSpeciesDict(self.parameters, self.multi_attempt_solution).to_dict(
                to_numpy=to_numpy, **kwargs
            )
        else:
            raise ValueError(f"Invalid output format: {output_format}")

    def compare(self, d1: dict, rtol: float, atol: float, log: bool = False) -> bool:
        """Compares a target dictionary to the model output.

        .. warning::
            Not compatible with JAX-compiled workflows (e.g., inside a :func:`jax.jit` context)

        Args:
            d1: Target dictionary
            rtol: Relative tolerance for comparison
            atol: Absolute tolerance for comparison
            log: Whether to compare the base-10 logarithm of the values. Defaults to ``False``.

        Returns:
            ``True`` if all values match within the specified tolerances, else ``False``
        """
        return OutputNamedArraysDict(self.parameters, self.multi_attempt_solution).compare(
            d1, rtol, atol, log
        )

    def quick_look(
        self,
        output_format: Literal["natural", "named_arrays", "elements_species"] = "named_arrays",
    ) -> None:
        """Quick look at the output.

        .. warning::
            Not compatible with JAX-compiled workflows (e.g., inside a :func:`jax.jit` context)

        Args:
            output_format: The format of the output dictionary. Can be ``natural`` for the natural
                output format based on the arrays used internally, ``named_arrays`` for an
                alternative format with named arrays, or ``elements_species`` for an alternative
                format grouped by element and species names. Defaults to ``named_arrays``.
        """
        out: dict[str, Any] = self.to_dict(output_format=output_format, to_numpy=True)
        logger.info("Quick look output:\n%s", pformat(out, sort_dicts=False))

    def _drop_unsuccessful_solves(
        self, dataframes: dict[str, pd.DataFrame]
    ) -> dict[str, pd.DataFrame]:
        """Drops unsuccessful solves.

        .. warning::
            Not compatible with JAX-compiled workflows (e.g., inside a :func:`jax.jit` context)

        Args:
            dataframes: Dataframes from which to drop unsuccessful models

        Returns:
            Dictionary of dataframes without unsuccessful models
        """
        return {
            key: df.loc[np.asarray(self.multi_attempt_solution.solver_success)]
            for key, df in dataframes.items()
        }

    def to_dataframes(
        self,
        output_format: Literal["named_arrays", "elements_species"] = "named_arrays",
        drop_unsuccessful_solves: bool = False,
    ) -> dict[str, pd.DataFrame]:
        """Gets the output in a dictionary of dataframes.

        Each top-level key becomes a DataFrame, with columns formed by joining nested keys with "."

        .. warning::
            Not compatible with JAX-compiled workflows (e.g., inside a :func:`jax.jit` context)

        Args:
            output_format: The format of the output dictionary. Can be ``natural`` for the natural
                output format based on the arrays used internally, ``named_arrays`` for an
                alternative format with named arrays, or ``elements_species`` for an alternative
                format grouped by element and species names. Defaults to ``named_arrays``.
            drop_unsuccessful_solves: Whether to drop unsuccessful solves from the output. Defaults
                to ``False``.

        Returns:
            Dictionary mapping top-level keys to pandas DataFrames
        """
        nested_dict: dict[str, Any] = self.to_dict(
            output_format=output_format, to_numpy=True, expand_to_batch=True, ravel=True
        )

        result: dict[str, pd.DataFrame] = {}

        for top_key, value in nested_dict.items():
            if isinstance(value, dict):
                flat: dict = flatten_dictionary(value)
                result[top_key] = pd.DataFrame(flat)
            else:
                result[top_key] = pd.DataFrame({top_key: value})

        if drop_unsuccessful_solves:
            logger.info("Dropping unsuccessful solves from output")
            result = self._drop_unsuccessful_solves(result)

        return result

    def to_excel(
        self,
        file_prefix: Path | str = "atmodeller_out",
        output_format: Literal["named_arrays", "elements_species"] = "named_arrays",
        drop_unsuccessful_solves: bool = False,
        append_version: bool = True,
    ) -> None:
        """Writes the output to an Excel file.

        .. warning::
            Not compatible with JAX-compiled workflows (e.g., inside a :func:`jax.jit` context)

        Args:
            file_prefix: Prefix of the output file. Accepts ``str`` or :class:`pathlib.Path`.
                Defaults to atmodeller_out.
            output_format: The format of the output dictionary. Can be ``natural`` for the natural
                output format based on the arrays used internally, ``named_arrays`` for an
                alternative format with named arrays, or ``elements_species`` for an alternative
                format grouped by element and species names. Defaults to ``named_arrays``.
            drop_unsuccessful_solves: Whether to drop unsuccessful solves from the output. Defaults
                to ``False``.
            append_version: Whether to append ``_v<package_version>`` to the output filename.
                Defaults to ``True``.
        """
        logger.info("Writing output to excel")
        out: dict[str, pd.DataFrame] = self.to_dataframes(
            output_format=output_format, drop_unsuccessful_solves=drop_unsuccessful_solves
        )
        output_path: Path = Path(file_prefix)
        base_name: str = output_path.name
        if base_name.endswith(".xlsx"):
            base_name = base_name[: -len(".xlsx")]
        if append_version:
            version_tag: str = f"_v{__version__}"
            if not base_name.endswith(version_tag):
                base_name = f"{base_name}{version_tag}"
        output_file: Path = output_path.with_name(f"{base_name}.xlsx")

        # Convenient to highlight rows where the solver failed to find a solution for follow-up
        # analysis. Define a fill colour for highlighting rows (e.g., yellow)
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Get the indices where the successful_solves mask is False
        unsuccessful_indices: NpArray = np.where(
            ~np.asarray(self.multi_attempt_solution.solver_success)
        )[0]

        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            for df_name, df in out.items():
                df.to_excel(writer, sheet_name=df_name, index=True)
                sheet = writer.sheets[df_name]

                # Apply highlighting to the rows where the solver failed to find a solution
                for idx in unsuccessful_indices:
                    # Highlight the entire row (starting from index 2 to skip header row)
                    for col in range(1, len(df.columns) + 2):
                        # row=idx+2 because Excel is 1-indexed and row 1 is the header
                        cell = sheet.cell(row=idx + 2, column=col)
                        cell.fill = highlight_fill

        logger.info("Output written to %s", output_file)

    def to_pickle(
        self,
        file_prefix: Path | str = "atmodeller_out",
        output_format: Literal["named_arrays", "elements_species"] = "named_arrays",
        drop_unsuccessful_solves: bool = False,
        append_version: bool = True,
    ) -> None:
        """Writes the output to a pickle file.

        .. warning::
            Not compatible with JAX-compiled workflows (e.g., inside a :func:`jax.jit` context)

        Args:
            file_prefix: Prefix of the output file. Accepts ``str`` or :class:`pathlib.Path`.
                Defaults to atmodeller_out.
            output_format: The format of the output dictionary. Can be ``natural`` for the natural
                output format based on the arrays used internally, ``named_arrays`` for an
                alternative format with named arrays, or ``elements_species`` for an alternative
                format grouped by element and species names. Defaults to ``named_arrays``.
            drop_unsuccessful_solves: Whether to drop unsuccessful solves from the output. Defaults
                to ``False``.
            append_version: Whether to append ``_v<package_version>`` to the output filename.
                Defaults to ``True``.
        """
        logger.info("Writing output to pickle")
        out: dict[str, pd.DataFrame] = self.to_dataframes(
            output_format=output_format, drop_unsuccessful_solves=drop_unsuccessful_solves
        )
        output_path: Path = Path(file_prefix)
        base_name: str = output_path.name
        if base_name.endswith(".pkl"):
            base_name = base_name[: -len(".pkl")]
        if append_version:
            version_tag: str = f"_v{__version__}"
            if not base_name.endswith(version_tag):
                base_name = f"{base_name}{version_tag}"
        output_file: Path = output_path.with_name(f"{base_name}.pkl")

        with open(output_file, "wb") as handle:
            pickle.dump(out, handle, protocol=pickle.HIGHEST_PROTOCOL)

        logger.info("Output written to %s", output_file)

    def solver_stats_to_logger(self) -> None:
        """Logs solver statistics.

        .. warning::
            Not compatible with JAX-compiled workflows (e.g., inside a :func:`jax.jit` context)
        """
        return self.multi_attempt_solution.stats_to_logger()
